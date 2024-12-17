import os
try:
    from openpyxl import Workbook
    import openpyxl
except ModuleNotFoundError:
    print("Installing openpyxl module")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl import Workbook
try:
    import matplotlib.pyplot as plt
except ModuleNotFoundError:
    print("Installing matplotlib module")
    os.system("pip install matplotlib")
    import matplotlib.pyplot as plt
try:
    import tabulate
except ModuleNotFoundError:
    print("Installing tabulate module")
    os.system("pip install tabulate")
    import tabulate
    
class User:
    def __init__(self, email, username, companyname):
        self.email = email
        self.username = username
        self.companyname = companyname
        self.input_data()

    def input_data(self):
        """Takes user input for energy, waste, and business travel."""
        try:                       
            # Energy Usage
            electricity_bill = float(input("Enter monthly electricity bill (in $): "))
            natural_gas_bill = float(input("Enter monthly natural gas bill (in $): "))
            fuel_bill = float(input("Enter monthly fuel bill (in $): "))
            energy_usage = electricity_bill*12*(0.0005) + natural_gas_bill*12*(0.0053) + fuel_bill*12*(2.32)

            # Waste
            total_waste = float(input("Enter total waste generated per month (in kg): "))
            recycling_percentage = float(input("Enter recycled waste/composted percentage (%): "))
            total_waste = total_waste*12*(0.57-recycling_percentage)

            # Business Travel
            kilometers_travelled = float(input("Enter total kilometers travelled for business purposes: "))
            fuel_efficiency = float(input("Enter average fuel efficiency (L/1000 km): "))
            business_travel = kilometers_travelled * (1/fuel_efficiency)*2.31

            # Combine data
            data = {
                "email": self.email,
                "username": self.username,
                "companyname": self.companyname,
                "energy_usage": round(energy_usage, 2),
                "total_waste": round(total_waste, 2),
                "business_travel": round(business_travel, 2),
            }
            self.save_to_excel(data)
        except ValueError:
            print("Invalid input! Please enter numeric values.")

    @staticmethod
    def save_to_excel(data):
        """Saves user data to an Excel file."""
        file_name = "UserData.xlsx"

        # Check if file exists, if not create a new one
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.append(["Email", "Username", "Company Name", "Carbon Emission (Energy Usage)", "Carbon Emission (Total Waste)", "Carbon Emission (Business Travel)"])
            wb.save(file_name)

        # Append data to the Excel file
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        ws.append([
            data["email"],
            data["username"],
            data["companyname"],
            data["energy_usage"],
            data["total_waste"],
            data["business_travel"],
        ])
        wb.save(file_name)
        print("Data saved successfully!")

    @staticmethod
    def show_data():
        """Displays data from the Excel file."""
        file_name = "UserData.xlsx"
    
        if not os.path.exists(file_name):
            print("No data available yet.")
            return
    
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
    
        print("\nUser Data:")
        data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = row
                continue
            data.append(row)
        print(tabulate.tabulate(data,headers=headers, tablefmt="fancy_grid"))

    
    @staticmethod
    def show_trends():
        """Displays trends using matplotlib charts."""
        file_name = "UserData.xlsx"

        if not os.path.exists(file_name):
            print("No data available yet.")
            return

        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

        usernames = []
        energy_usage = []
        total_waste = []
        business_travel = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            usernames.append(row[1])  # Username
            energy_usage.append(row[3])  # Energy Usage
            total_waste.append(row[4])  # Total Waste
            business_travel.append(row[5])  # Business Travel

        # Prompt the user for the desired parameter to plot
        while True:
            print("Select a parameter to view trends:")
            print("1. Energy Usage")
            print("2. Total Waste")
            print("3. Business Travel")
            print("4. Overall Carbon Emissions (Pie Chart)")
            print("5. Back")
            choice = input("Enter the number of your choice: ")

            if choice == "1":
                data = energy_usage
                label = "Energy Usage"
            elif choice == "2":
                data = total_waste
                label = "Total Waste"
            elif choice == "3":
                data = business_travel
                label = "Business Travel"
            elif choice == "4":
                # Display pie chart for overall emissions
                total_energy = sum(energy_usage)
                total_waste_emissions = sum(total_waste)
                total_travel = sum(business_travel)
                category_totals = {
                    "Energy": total_energy,
                    "Waste": total_waste_emissions,
                    "Travel": total_travel
                }

                plt.figure(figsize=(8, 8))
                plt.pie(
                    category_totals.values(),
                    labels=category_totals.keys(),
                    autopct='%1.1f%%',
                    startangle=140,
                    colors=['lightblue', 'lightgreen', 'salmon']
                )
                plt.title("Overall Carbon Emissions by Category")
                plt.show()
                continue
            elif choice == "5":
                return
            else:
                print("Invalid choice!")
                return

            # Plotting the trend
            plt.figure(figsize=(10, 6))
            plt.plot(usernames, data, marker='o', label=label)

            # Highlight the highest value
            max_value = max(data)
            max_index = data.index(max_value)
            plt.plot(usernames[max_index], max_value, marker='o', markersize=10, color='red', label="Highest Emission")

            plt.xlabel('Users')
            plt.ylabel('Values')
            plt.title(f'{label} Trend')
            plt.xticks(rotation=45)
            plt.legend()
            plt.tight_layout()
            plt.show()
    
# Main Driver Function starts here
if __name__ == "__main__":
    
    print("Welcome to the Carbon Footprint Tracker")

    email = input("Enter your email: ")
    username = input("Enter your username: ")
    companyname = input("Enter your company name: ")
    user = User(email, username, companyname)
    
    while True:
        print("\nMenu:")
        print("1. Input Data")
        print("2. Show Data")
        print("3. Show Trends")
        print("4. Exit")

        choice = input("Enter your choice: ")
            
        if choice == "1":
            user.email = input("Enter your email: ")
            user.username = input("Enter your username: ")
            user.companyname = input("Enter your company name: ")
            user.input_data()
        elif choice == "2":
            User.show_data()
        elif choice == "3":
            User.show_trends()
        elif choice == "4":
            print("Goodbye!")
            break
        else:
            print("Invalid choice. Please try again.")
